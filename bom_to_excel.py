#!/usr/bin/env python3
"""Convert a costed BOM PDF to an editable Excel workbook (CURR columns only)."""

import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

UOM_PATTERN = re.compile(
    r"^(EA|M3|GA|RL|FT|PC|BOOT|PCS|LB|KG|IN|SF|SY|LF|CY|TON|GAL|QT|PT|OZ|ML|L|MM|CM|M)$"
)
NUM_OR_STAR = re.compile(r"^(\d+\.?\d*|\*{8})$")


def parse_value(token: str) -> float | None:
    if token == "********":
        return None
    return float(token)


def parse_line_item(line: str) -> dict | None:
    """Parse a BOM line item; returns None if not a line item."""
    line = line.strip()
    if not line or line.startswith("Desc:") or line.startswith("--------"):
        return None

    if line.startswith("TOTAL "):
        parts = line.split()
        if (
            len(parts) >= 5
            and NUM_OR_STAR.match(parts[-4])
            and NUM_OR_STAR.match(parts[-3])
        ):
            return {
                "type": "section_total",
                "section": " ".join(parts[1:-4]),
                "curr_cost": parse_value(parts[-4]),
                "curr_waste": parse_value(parts[-3]),
            }
        return None

    if not re.match(r"^\d{5}\s", line):
        return None

    tokens = line.split()
    if len(tokens) < 9:
        return None

    tail = tokens[-8:]
    if not all(NUM_OR_STAR.match(t) for t in tail[2:]):
        return None
    if not UOM_PATTERN.match(tail[1]):
        return None

    head = tokens[:-8]
    return {
        "type": "item",
        "part": head[0],
        "description": " ".join(head[1:]),
        "qty": float(tail[0]),
        "uom": tail[1],
        "curr_per": parse_value(tail[2]),
        "curr_cost": parse_value(tail[3]),
        "curr_waste": parse_value(tail[4]),
    }


def extract_metadata(text: str) -> dict:
    meta = {}
    for key, pattern in [
        ("item_no", r"ITEM NO\s*:\s*(\S+)"),
        ("entity", r"ENTITY\s*:\s*(\S+)"),
        ("warehouse", r"WAREHOUSE\s*:\s*(\S+)"),
        ("generic", r"Generic\s*:\s*(\S+)"),
        ("date", r"(\d{1,2}/\d{1,2}/\d{2})\s+\d{2}:\d{2}"),
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            meta[key] = clean_text(m.group(1))

    desc_lines = []
    for line in text.splitlines():
        if line.strip().startswith("Desc:"):
            desc_lines.append(line.strip()[5:].strip())
        elif desc_lines and not re.match(r"^(CHEM|HDWR|\d{5})", line.strip()):
            if line.strip() and not line.strip().startswith("PART#"):
                desc_lines.append(line.strip())
            else:
                break
        elif desc_lines:
            break
    if desc_lines:
        meta["description"] = " ".join(desc_lines)
    return meta


def parse_summary(text: str) -> list[dict]:
    """Extract summary rows from the detailed totals page (CURR cost only)."""
    match = re.search(r"Totals for Finished Item:.*", text, re.DOTALL)
    block = match.group(0) if match else text
    rows = []
    patterns = [
        (r"^MATERIAL\s+([\d.]+)", "MATERIAL", "material"),
        (r"^VEND-FRT & DISC\s+([\d.]+)", "VEND-FRT & DISC", "editable"),
        (r"^MISC MATERIAL\s+([\d.]+)", "MISC MATERIAL", "editable"),
        (r"^PALLET MATERIAL\s+([\d.]+)", "PALLET MATERIAL", "editable"),
        (r"^TOTAL MISC & PALLET\s+([\d.]+)", "TOTAL MISC & PALLET", "sum_misc_pallet"),
        (r"^WASTE\s+([\d.]+)", "WASTE", "waste"),
        (
            r"^TOTAL MATERIAL\([^)]*\)\s+([\d.]+)",
            "TOTAL MATERIAL",
            "sum_material_block",
        ),
        (r"^LABOR FROM ROUTINGS?\s+([\d.]+)", "LABOR FROM ROUTING", "editable"),
        (r"^RECEIVING\([^)]*\)\s+([\d.]+)", "RECEIVING", "editable"),
        (r"^WAREHOUSING\([^)]*\)\s+([\d.]+)", "WAREHOUSING", "editable"),
        (r"^TOTAL LABOR\([^)]*\)\s+([\d.]+)", "TOTAL LABOR", "sum_labor"),
        (r"^TAXES & BENEFITS\([^)]*\)\s+([\d.]+)", "TAXES & BENEFITS", "editable"),
        (
            r"^TOTAL LABOR, TAXES & BENEFITS\s+([\d.]+)",
            "TOTAL LABOR, TAXES & BENEFITS",
            "sum_labor_block",
        ),
        (r"^TOTAL INVENTORY COST\s+([\d.]+)", "TOTAL INVENTORY COST", "sum_inventory"),
        (r"^FREIGHT\([^)]*\)\s+([\d.]+)", "FREIGHT", "editable"),
        (r"^SALES ALLOWANCE\([^)]*\)\s+([\d.]+)", "SALES ALLOWANCE", "editable"),
        (r"^TOTAL COST\s+([\d.]+)", "TOTAL COST", "total_cost"),
    ]

    for line in block.splitlines():
        line = re.sub(r"\(cid:\d+\)", "", line).strip()
        for pattern, label, kind in patterns:
            m = re.match(pattern, line)
            if m:
                rows.append({"label": label, "kind": kind, "value": float(m.group(1))})
                break
    return rows


def style_header_row(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_section_row(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="D6E4F0")
    font = Font(bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font


def style_total_row(ws, row: int, cols: int) -> None:
    font = Font(bold=True)
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).font = font


def set_currency(ws, row: int, cols: list[int]) -> None:
    for col in cols:
        ws.cell(row=row, column=col).number_format = "#,##0.0000"


def clean_text(value: str) -> str:
    return re.sub(r"\(cid:\d+\)", "", value).strip()


def write_summary_panel(
    ws,
    summary: list[dict],
    section_total_rows: list[int],
    start_col: int,
    start_row: int,
) -> None:
    """Place cost summary in a right-side panel (beside line items, not below)."""
    label_col = start_col
    value_col = start_col + 1
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=start_row, column=label_col, value="COST SUMMARY").font = Font(
        bold=True, size=12
    )
    ws.merge_cells(
        start_row=start_row,
        start_column=label_col,
        end_row=start_row,
        end_column=value_col,
    )

    header_row = start_row + 1
    ws.cell(row=header_row, column=label_col, value="Line Item")
    ws.cell(row=header_row, column=value_col, value="Amount")
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(label_col, value_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    sum_row = header_row + 1
    summary_row_map: dict[str, int] = {}
    value_letter = get_column_letter(value_col)

    material_formula = (
        f"=SUM({','.join(f'G{r}' for r in section_total_rows)})"
        if section_total_rows
        else "=0"
    )
    waste_formula = (
        f"=SUM({','.join(f'H{r}' for r in section_total_rows)})"
        if section_total_rows
        else "=0"
    )

    for entry in summary:
        label = entry["label"]
        kind = entry["kind"]
        ws.cell(row=sum_row, column=label_col, value=label)

        if kind == "material":
            ws.cell(row=sum_row, column=value_col, value=material_formula)
        elif kind == "waste":
            ws.cell(row=sum_row, column=value_col, value=waste_formula)
        elif kind == "sum_misc_pallet":
            misc = summary_row_map["MISC MATERIAL"]
            pallet = summary_row_map["PALLET MATERIAL"]
            ws.cell(row=sum_row, column=value_col, value=f"={value_letter}{misc}+{value_letter}{pallet}")
        elif kind == "sum_material_block":
            refs = [
                summary_row_map["MATERIAL"],
                summary_row_map["VEND-FRT & DISC"],
                summary_row_map["TOTAL MISC & PALLET"],
                summary_row_map["WASTE"],
            ]
            ws.cell(
                row=sum_row,
                column=value_col,
                value="=" + "+".join(f"{value_letter}{r}" for r in refs),
            )
        elif kind == "sum_labor":
            refs = [
                summary_row_map["LABOR FROM ROUTING"],
                summary_row_map["RECEIVING"],
                summary_row_map["WAREHOUSING"],
            ]
            ws.cell(
                row=sum_row,
                column=value_col,
                value="=" + "+".join(f"{value_letter}{r}" for r in refs),
            )
        elif kind == "sum_labor_block":
            refs = [
                summary_row_map["TOTAL LABOR"],
                summary_row_map["TAXES & BENEFITS"],
            ]
            ws.cell(
                row=sum_row,
                column=value_col,
                value=f"={value_letter}{refs[0]}+{value_letter}{refs[1]}",
            )
        elif kind == "sum_inventory":
            refs = [
                summary_row_map["TOTAL MATERIAL"],
                summary_row_map["TOTAL LABOR, TAXES & BENEFITS"],
            ]
            ws.cell(
                row=sum_row,
                column=value_col,
                value=f"={value_letter}{refs[0]}+{value_letter}{refs[1]}",
            )
        elif kind == "total_cost":
            refs = [
                summary_row_map["TOTAL INVENTORY COST"],
                summary_row_map["FREIGHT"],
                summary_row_map["SALES ALLOWANCE"],
            ]
            ws.cell(
                row=sum_row,
                column=value_col,
                value=f"={value_letter}{refs[0]}+{value_letter}{refs[1]}+{value_letter}{refs[2]}",
            )
            ws.cell(row=sum_row, column=label_col).font = Font(bold=True, size=12)
            ws.cell(row=sum_row, column=value_col).font = Font(bold=True, size=12)
            fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row=sum_row, column=label_col).fill = fill
            ws.cell(row=sum_row, column=value_col).fill = fill
        else:
            ws.cell(row=sum_row, column=value_col, value=entry["value"])

        set_currency(ws, sum_row, [value_col])
        for col in range(label_col, value_col + 1):
            ws.cell(row=sum_row, column=col).border = border
        if kind == "total_cost":
            style_total_row(ws, sum_row, value_col - label_col + 1)
        summary_row_map[label] = sum_row
        sum_row += 1


def page_has_line_items(text: str) -> bool:
    return bool(re.search(r"^\d{5}\s", text, re.MULTILINE))


def page_starts_bom(text: str) -> bool:
    return (
        "COSTED BILL OF MATERIAL" in text
        and "PART#" in text
        and page_has_line_items(text)
    )


def split_into_bom_blocks(page_texts: list[str]) -> list[dict]:
    blocks: list[dict] = []
    item_pages: list[str] = []
    bom_text_parts: list[str] = []

    def push_block(summary_text: str | None) -> None:
        nonlocal item_pages, bom_text_parts
        if not item_pages:
            return

        parts = [*bom_text_parts]
        if summary_text:
            parts.append(summary_text)
        full_text = "\n".join(parts)
        totals_match = re.search(r"Totals for Finished Item:\s*(\S+)", summary_text or "")
        item_no = clean_text(totals_match.group(1)) if totals_match else ""
        if not item_no:
            item_match = re.search(r"ITEM NO\s*:\s*(\S+)", full_text, re.IGNORECASE)
            item_no = clean_text(item_match.group(1)) if item_match else f"BOM {len(blocks) + 1}"

        blocks.append(
            {
                "item_no": item_no,
                "item_pages": list(item_pages),
                "summary_text": summary_text,
                "full_text": full_text,
            }
        )
        item_pages = []
        bom_text_parts = []

    for page_text in page_texts:
        totals_match = re.search(r"Totals for Finished Item:\s*(\S+)", page_text)
        has_line_items = page_has_line_items(page_text)

        if page_starts_bom(page_text) and item_pages:
            push_block(None)

        if has_line_items:
            item_pages.append(page_text)
            bom_text_parts.append(page_text)
        elif "COSTED BILL OF MATERIAL" in page_text and "ITEM NO" in page_text:
            bom_text_parts.append(page_text)

        if totals_match:
            push_block(page_text)

    if item_pages:
        push_block(None)

    return blocks


def parse_sections_from_pages(page_texts: list[str]) -> list[tuple[str, list]]:
    sections: list[tuple[str, list]] = []
    current_section = None

    for page_text in page_texts:
        for raw_line in page_text.splitlines():
            line = clean_text(raw_line)
            if not line:
                continue

            if re.match(r"^[A-Z][A-Z0-9/-]*$", line) and not line.startswith("Desc"):
                if line not in ("PAGE",) and len(line) < 20:
                    current_section = line
                    sections.append((current_section, []))
                continue

            parsed = parse_line_item(line)
            if parsed and current_section:
                sections[-1][1].append(parsed)

    return sections


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name).strip()[:31]
    return cleaned or "BOM"


def unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    candidate = sanitize_sheet_name(base_name)
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    index = 2
    while index < 100:
        suffix = f" ({index})"
        candidate = sanitize_sheet_name(f"{base_name[: 31 - len(suffix)]}{suffix}")
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1

    used_names.add(candidate)
    return candidate


def write_bom_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: dict,
    sections: list[tuple[str, list]],
    summary: list[dict],
) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "COSTED BILL OF MATERIAL"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    meta_pairs = [
        ("Item No", clean_text(meta.get("item_no", ""))),
        ("Entity", clean_text(meta.get("entity", ""))),
        ("Warehouse", clean_text(meta.get("warehouse", ""))),
        ("Generic", clean_text(meta.get("generic", ""))),
        ("Date", clean_text(meta.get("date", ""))),
    ]
    col = 1
    for label, value in meta_pairs:
        ws.cell(row=2, column=col, value=label).font = Font(bold=True)
        ws.cell(row=2, column=col + 1, value=value)
        col += 2

    if meta.get("description"):
        ws.cell(row=3, column=1, value="Description").font = Font(bold=True)
        ws.cell(row=3, column=2, value=clean_text(meta["description"]))
        ws.merge_cells("B3:H3")
        ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
        header_row = 5
    else:
        header_row = 4

    headers = [
        "Category",
        "Part #",
        "Description",
        "Qty",
        "UOM",
        "Unit Cost",
        "Extended Cost",
        "Waste",
    ]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_index, value=header)
    style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    section_total_rows: list[int] = []

    for section_name, items in sections:
        if not items:
            continue

        item_rows: list[int] = []
        for item in items:
            if item["type"] == "section_total":
                continue

            ws.cell(row=row, column=1, value=section_name)
            ws.cell(row=row, column=2, value=item["part"])
            ws.cell(row=row, column=3, value=item["description"])
            ws.cell(row=row, column=4, value=item["qty"])
            ws.cell(row=row, column=5, value=item["uom"])
            if item["curr_per"] is not None:
                ws.cell(row=row, column=6, value=item["curr_per"])
            ws.cell(row=row, column=7, value=f"=D{row}*F{row}")
            if item["curr_waste"] is not None:
                ws.cell(row=row, column=8, value=item["curr_waste"])

            set_currency(ws, row, [4, 6, 7, 8])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).border = border
            item_rows.append(row)
            row += 1

        if item_rows:
            style_total_row(ws, row, len(headers))
            ws.cell(row=row, column=1, value=section_name)
            ws.cell(row=row, column=2, value=f"TOTAL {section_name}")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=7, value=f"=SUM(G{min(item_rows)}:G{max(item_rows)})")
            ws.cell(row=row, column=8, value=f"=SUM(H{min(item_rows)}:H{max(item_rows)})")
            set_currency(ws, row, [7, 8])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).border = border
                if col_idx == 1:
                    ws.cell(row=row, column=col_idx).font = Font(bold=True)
            section_total_rows.append(row)
            row += 1

    summary_col = 10
    write_summary_panel(ws, summary, section_total_rows, summary_col, header_row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 2
    ws.column_dimensions[get_column_letter(summary_col)].width = 28
    ws.column_dimensions[get_column_letter(summary_col + 1)].width = 14

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def convert_bom(pdf_path: Path, xlsx_path: Path) -> None:
    page_texts: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_texts.append(page.extract_text() or "")

    blocks = split_into_bom_blocks(page_texts)
    if not blocks:
        blocks = [
            {
                "item_no": "BOM",
                "item_pages": page_texts[:2],
                "summary_text": None,
                "full_text": "\n".join(page_texts),
            }
        ]

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    for block in blocks:
        full_text = block["full_text"] or "\n".join(block["item_pages"])
        meta = extract_metadata(full_text)
        meta["item_no"] = meta.get("item_no") or block["item_no"]
        sections = parse_sections_from_pages(block["item_pages"])
        summary = parse_summary(full_text)
        sheet_name = unique_sheet_name(block["item_no"], used_names)
        write_bom_sheet(wb, sheet_name, meta, sections, summary)

    wb.save(xlsx_path)
    sheet_count = len(wb.sheetnames)
    if sheet_count > 1:
        print(f"Created {xlsx_path} with {sheet_count} BOM sheets")
    else:
        print(f"Created {xlsx_path}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python bom_to_excel.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1]).expanduser().resolve()
    if len(sys.argv) >= 3:
        xlsx_path = Path(sys.argv[2]).expanduser().resolve()
    else:
        xlsx_path = pdf_path.with_suffix(".xlsx")

    convert_bom(pdf_path, xlsx_path)


if __name__ == "__main__":
    main()
